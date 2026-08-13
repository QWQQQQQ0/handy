"""Sandbox code execution handlers.

Extracted from main.py to keep the dispatcher file manageable.
Three handlers: exec_python (restricted), web_code_exec (Playwright),
and doc_code_exec (Office document libraries).

These use delayed imports from main.py (via `from main import X`) to
avoid circular imports — the handlers are only called during the
main() event loop, at which point main.py is fully loaded.
"""

from __future__ import annotations

from typing import Any


def _handle_exec_python(params: dict) -> dict:
    """Execute arbitrary Python code in a restricted sandbox.

    Allowed modules (SAFE_MODULES):
      json, math, datetime, re, collections, itertools, random, statistics,
      uuid, base64, hashlib, textwrap, string, typing, enum, functools,
      operator, bisect, decimal, fractions, copy.

    The value of a variable named ``result`` in the executed scope is
    returned as the ``result`` field.
    """
    code = params.get("code", "")
    timeout_sec = params.get("timeout_sec", 30)
    input_vars = params.get("params", {})
    allow_all = params.get("allowAllImports", False)

    import io
    import sys
    import traceback
    import time

    old_stdout, old_stderr = sys.stdout, sys.stderr
    captured_stdout, captured_stderr = io.StringIO(), io.StringIO()
    sys.stdout, sys.stderr = captured_stdout, captured_stderr

    SAFE_MODULES = {
        "json", "math", "datetime", "re", "collections", "itertools",
        "random", "statistics", "uuid", "base64", "hashlib", "textwrap",
        "string", "typing", "enum", "functools", "operator", "bisect",
        "decimal", "fractions", "copy",
    }

    def safe_import(name, *args):
        if name not in SAFE_MODULES:
            raise ImportError(f"Module '{name}' is not allowed")
        return __import__(name, *args)

    safe_globals = {
        "__builtins__": {
            "print": print,
            "len": len,
            "range": range,
            "int": int,
            "float": float,
            "str": str,
            "bool": bool,
            "list": list,
            "dict": dict,
            "tuple": tuple,
            "set": set,
            "sorted": sorted,
            "reversed": reversed,
            "enumerate": enumerate,
            "zip": zip,
            "map": map,
            "filter": filter,
            "any": any,
            "all": all,
            "min": min,
            "max": max,
            "sum": sum,
            "abs": abs,
            "round": round,
            "isinstance": isinstance,
            "type": type,
            "hasattr": hasattr,
            "getattr": getattr,
            "setattr": setattr,
            "ValueError": ValueError,
            "TypeError": TypeError,
            "KeyError": KeyError,
            "IndexError": IndexError,
            "Exception": Exception,
            "StopIteration": StopIteration,
            "__import__": __import__ if allow_all else safe_import,
        }
    }
    safe_globals.update(input_vars)

    result_value, error_str = None, None
    start = time.time()

    try:
        compiled = compile(code, "<sandbox>", "exec", flags=0)
        exec(compiled, safe_globals)
        result_value = safe_globals.get("result", None)
    except Exception:
        error_str = traceback.format_exc()
    finally:
        sys.stdout, sys.stderr = old_stdout, old_stderr

    duration_ms = int((time.time() - start) * 1000)
    output = captured_stdout.getvalue()
    err_output = captured_stderr.getvalue()

    return {
        "success": error_str is None,
        "output": output,
        "error": error_str or err_output or "",
        "result": result_value,
        "duration_ms": duration_ms,
        "truncated": len(output) > 100000,
    }


def _handle_web_code_exec(params: dict) -> dict:
    """Execute Python code with Playwright browser context injected.

    Pre-injected variables:
      - page: Playwright Page object (current active page)
      - browser: Playwright Browser object
      - wait_for: convenience wrapper for page.wait_for_selector
      - screenshot_b64(): take screenshot and return base64 string
      - navigate/click/fill/scroll/get_interactive/get_content/evaluate: helper functions

    The value of a variable named ``result`` is returned as the ``result`` field.
    """
    code = params.get("code", "")
    timeout_sec = params.get("timeout_sec", 60)

    import io
    import sys
    import time
    import traceback
    import base64

    # Delayed import — main.py is fully loaded by the time this handler runs.
    from main import _get_browser as _get_br
    browser_engine = _get_br()

    # Check if we have Playwright page OR extension connection
    has_playwright = browser_engine._page is not None
    has_extension = False
    try:
        from engine.extension_ws import is_extension_connected
        has_extension = is_extension_connected()
    except ImportError:
        pass

    if not has_playwright and not has_extension:
        return {
            "success": False,
            "output": "",
            "error": "No browser available. Connect browser extension or call web_launch first.",
            "result": None,
            "duration_ms": 0,
            "truncated": False,
        }

    page = browser_engine._page if has_playwright else None
    browser = browser_engine._browser if has_playwright else None

    # ── Dangerous operation guard ──
    _DANGEROUS_PATTERNS = [
        "os._exit", "os.system", "os.exec", "os.spawn",
        "os.kill", "os.remove", "os.unlink", "os.rmdir",
        "shutil.rmtree", "shutil.move",
        "subprocess.", "ctypes.",
        "importlib.",
    ]
    for pat in _DANGEROUS_PATTERNS:
        if pat in code:
            return {
                "success": False,
                "output": "",
                "error": f"Blocked dangerous operation: '{pat}' is not allowed in web_code_exec.",
                "result": None,
                "duration_ms": 0,
                "truncated": False,
            }

    # ── Capture stdout/stderr ──
    old_stdout, old_stderr = sys.stdout, sys.stderr
    captured_stdout, captured_stderr = io.StringIO(), io.StringIO()
    sys.stdout, sys.stderr = captured_stdout, captured_stderr

    # ── Extension helper (when Playwright is not available) ──
    def _ext_exec(js_code):
        """Execute JavaScript via extension and return result."""
        from engine.extension_ws import _handle_command
        result = _handle_command("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return result["data"]["results"][0] if result["data"]["results"] else None
        raise Exception(result.get("error", "Extension execution failed"))

    # ── Helper functions (work with both Playwright and extension) ──
    def screenshot_b64():
        """Take a screenshot and return as base64 string."""
        if has_playwright:
            img_bytes = page.screenshot()
            return base64.b64encode(img_bytes).decode("ascii")
        raise Exception("Screenshot requires Playwright browser. Call web_launch first.")

    def wait_for(selector, timeout=10000):
        """Wait for selector to appear, return the element."""
        if has_playwright:
            return page.wait_for_selector(selector, timeout=timeout)
        # Extension mode: poll with JavaScript
        import time as _time
        deadline = _time.time() + timeout / 1000
        while _time.time() < deadline:
            exists = _ext_exec(f"!!document.querySelector('{selector}')")
            if exists:
                return True
            _time.sleep(0.5)
        raise TimeoutError(f"Selector '{selector}' not found within {timeout}ms")

    def navigate(url, wait_until="load"):
        """Navigate to URL."""
        if has_playwright:
            page.goto(url, wait_until=wait_until)
            return {"url": page.url, "title": page.title()}
        _ext_exec(f"window.location.href = '{url}'")
        return {"url": url, "navigated": True}

    def click(selector):
        """Click element by selector."""
        if has_playwright:
            page.click(selector)
            return {"clicked": selector}
        _ext_exec(f"document.querySelector('{selector}').click()")
        return {"clicked": selector}

    def fill(selector, value):
        """Fill input field."""
        if has_playwright:
            page.fill(selector, value)
            return {"filled": selector}
        escaped = value.replace("'", "\\'").replace("\n", "\\n")
        _ext_exec(f"document.querySelector('{selector}').value = '{escaped}'")
        return {"filled": selector}

    def scroll(direction="down", amount=500):
        """Scroll page. direction: 'up', 'down', 'left', 'right'."""
        js = {
            "down": f"window.scrollBy(0, {amount})",
            "up": f"window.scrollBy(0, -{amount})",
            "right": f"window.scrollBy({amount}, 0)",
            "left": f"window.scrollBy(-{amount}, 0)",
        }.get(direction, f"window.scrollBy(0, {amount})")
        if has_playwright:
            page.evaluate(js)
        else:
            _ext_exec(js)
        return {"scrolled": direction, "amount": amount}

    def get_interactive():
        """Get all interactive elements on the page with selectors."""
        js = """() => {
            const elements = document.querySelectorAll('a, button, input, select, textarea, [role="button"], [onclick]');
            return Array.from(elements).map(el => ({
                tag: el.tagName.toLowerCase(),
                text: el.textContent?.trim()?.substring(0, 100) || '',
                id: el.id || '',
                name: el.name || '',
                type: el.type || '',
                href: el.href || '',
                placeholder: el.placeholder || '',
                value: el.value || '',
                selector: el.id ? `#${el.id}` : el.name ? `[name="${el.name}"]` : null
            })).filter(e => e.text || e.id || e.name || e.placeholder);
        }"""
        if has_playwright:
            return page.evaluate(js)
        return _ext_exec(js)

    def get_content(text_only=True, max_length=50000):
        """Get page content. Set text_only=False for HTML."""
        if has_playwright:
            if text_only:
                content = page.evaluate("document.body.innerText")
            else:
                content = page.content()
        else:
            js = "document.body.innerText" if text_only else "document.documentElement.outerHTML"
            content = _ext_exec(js)
        return content[:max_length] if len(content) > max_length else content

    def evaluate(js_code):
        """Execute JavaScript in the page context and return result."""
        if has_playwright:
            return page.evaluate(js_code)
        return _ext_exec(js_code)

    def close_browser():
        """Close the browser."""
        if has_playwright:
            browser.close()
            return {"closed": True}
        return {"closed": False, "reason": "Extension mode - cannot close browser"}

    # ── Build sandbox globals ──
    safe_globals = {
        "__builtins__": {
            "print": print,
            "len": len,
            "range": range,
            "int": int,
            "float": float,
            "str": str,
            "bool": bool,
            "list": list,
            "dict": dict,
            "tuple": tuple,
            "set": set,
            "sorted": sorted,
            "reversed": reversed,
            "enumerate": enumerate,
            "zip": zip,
            "map": map,
            "filter": filter,
            "any": any,
            "all": all,
            "min": min,
            "max": max,
            "sum": sum,
            "abs": abs,
            "round": round,
            "isinstance": isinstance,
            "type": type,
            "hasattr": hasattr,
            "getattr": getattr,
            "setattr": setattr,
            "ValueError": ValueError,
            "TypeError": TypeError,
            "KeyError": KeyError,
            "IndexError": IndexError,
            "Exception": Exception,
            "StopIteration": StopIteration,
            "__import__": lambda name, *args: __import__(name, *args) if name in {"time", "json", "re", "base64", "math"} else (_ for _ in ()).throw(ImportError(f"Module '{name}' is not allowed")),
        },
        "page": page,
        "browser": browser,
        "wait_for": wait_for,
        "screenshot_b64": screenshot_b64,
        # Web operation helpers
        "navigate": navigate,
        "click": click,
        "fill": fill,
        "scroll": scroll,
        "get_interactive": get_interactive,
        "get_content": get_content,
        "evaluate": evaluate,
        "close_browser": close_browser,
    }

    result_value, error_str = None, None
    start = time.time()

    try:
        # Execute with timeout
        import threading

        def run_code():
            nonlocal result_value, error_str
            try:
                compiled = compile(code, "<web_sandbox>", "exec", flags=0)
                exec(compiled, safe_globals)
                result_value = safe_globals.get("result", None)
            except Exception:
                error_str = traceback.format_exc()

        thread = threading.Thread(target=run_code)
        thread.start()
        thread.join(timeout=timeout_sec)

        if thread.is_alive():
            error_str = f"Code execution timed out after {timeout_sec}s"
    except Exception:
        error_str = traceback.format_exc()
    finally:
        sys.stdout, sys.stderr = old_stdout, old_stderr

    duration_ms = int((time.time() - start) * 1000)
    output = captured_stdout.getvalue()
    err_output = captured_stderr.getvalue()

    return {
        "success": error_str is None,
        "output": output,
        "error": error_str or err_output or "",
        "result": result_value,
        "duration_ms": duration_ms,
        "truncated": len(output) > 100000,
    }


def _handle_doc_code_exec(params: dict) -> dict:
    """Execute Python code with full access to document libraries and COM instances.

    Pre-injected variables available in user code:
      - openpyxl: Workbook, load_workbook
      - docx: Document (python-docx)
      - pptx: Presentation (python-pptx)
      - ExcelCOM, WordCOM, PptCOM: COM classes
      - ExcelGenerator, WordGenerator, PptGenerator: file generators
      - get_excel_app(): get or create ExcelCOM instance (connects to user's open Excel)
      - get_word_app(): get or create WordCOM instance
      - get_ppt_app(): get or create PptCOM instance
      - read_range(addr, sheet=None, file_path=None): convenience for ExcelCOM.read_range
      - save_workbook(): convenience for ExcelCOM.save

    The value of a variable named ``result`` is returned as the ``result`` field.
    """
    code = params.get("code", "")
    timeout_sec = params.get("timeout_sec", 60)

    import io
    import sys
    import time
    import threading
    import json
    from datetime import datetime

    # ── Delayed imports from main.py ──
    from main import _get_excel_com, _get_word_com, _get_ppt_com
    from main import _get_excel_gen, _get_word_gen, _get_ppt_gen

    # ── COM classes (optional — may be None if pywin32 missing) ──
    try:
        from engine.office import WordCOM, ExcelCOM, PptCOM
    except ImportError:
        WordCOM = None  # type: ignore[assignment,misc]
        ExcelCOM = None  # type: ignore[assignment,misc]
        PptCOM = None  # type: ignore[assignment,misc]

    # ── Generator classes ──
    from engine.office import WordGenerator, ExcelGenerator, PptGenerator

    # ── Dangerous operation guard ──
    _DANGEROUS_PATTERNS = [
        # Process / system
        "os._exit", "os.system", "os.exec", "os.spawn", "os.popen",
        "os.kill", "os.remove", "os.unlink", "os.rmdir",
        "os.rename", "os.chmod", "os.chown", "os.symlink",
        "shutil.rmtree", "shutil.move", "shutil.copy", "shutil.copytree",
        "subprocess.", "ctypes.",
        "importlib.", "pkgutil.",
        # Networking (docs don't need it)
        "socket.", "urllib.request", "urllib.error",
        "http.client", "http.server",
        "requests.get", "requests.post",
        "ftplib.", "smtplib.",
        # Code execution
        "exec(", "eval(", "compile(",
        "__builtins__",
        "globals()", "locals()",
        # Registry (Windows)
        "winreg.", "win32api.Reg",
        # Threading abuse
        "multiprocessing.",
        # Dynamic import evasion
        "__import__(",
    ]
    for pat in _DANGEROUS_PATTERNS:
        if pat in code:
            return {
                "success": False,
                "output": "",
                "error": f"Blocked dangerous operation: '{pat}' is not allowed in doc_code_exec. "
                         f"Use the predefined com_read/com_edit tools for file operations, "
                         f"or doc_code_exec for document data processing only.",
                "result": None,
                "duration_ms": 0,
                "truncated": False,
            }

    # ── Volume limit guard ──
    if len(code) > 20000:
        return {
            "success": False,
            "output": "",
            "error": "Code exceeds maximum length of 20000 characters.",
            "result": None,
            "duration_ms": 0,
            "truncated": False,
        }

    # ── Capture stdout/stderr ──
    old_stdout, old_stderr = sys.stdout, sys.stderr
    captured_stdout, captured_stderr = io.StringIO(), io.StringIO()
    sys.stdout, sys.stderr = captured_stdout, captured_stderr

    # ── Build sandbox globals with document libraries ──
    try:
        from openpyxl import Workbook as _OWB, load_workbook as _load_wb
        from openpyxl.styles import Font as _Font, Alignment as _Align, PatternFill as _Fill, Border as _Border, Side as _Side
    except ImportError:
        _OWB = None  # type: ignore
        _load_wb = None  # type: ignore
        _Font = _Align = _Fill = _Border = _Side = None

    try:
        from docx import Document as _DocxDoc
    except ImportError:
        _DocxDoc = None  # type: ignore

    try:
        from pptx import Presentation as _PptxPres
    except ImportError:
        _PptxPres = None  # type: ignore

    # Convenience: lazy COM instance getters
    _com_cache: dict = {}

    def get_excel_app():
        if "excel" not in _com_cache:
            _com_cache["excel"] = _get_excel_com()
        return _com_cache["excel"]

    def get_word_app():
        if "word" not in _com_cache:
            _com_cache["word"] = _get_word_com()
        return _com_cache["word"]

    def get_ppt_app():
        if "ppt" not in _com_cache:
            _com_cache["ppt"] = _get_ppt_com()
        return _com_cache["ppt"]

    def read_range(addr, sheet=None, file_path=None):
        return get_excel_app().read_range(addr, sheet=sheet, file_path=file_path)

    def save_workbook():
        return get_excel_app().save()

    def save_document():
        return get_word_app().save()

    def save_presentation():
        return get_ppt_app().save()

    # Document detection (replaces office_detect tool)
    try:
        from engine.office.com_resolver import detect_all as _detect_all
    except ImportError:
        _detect_all = None  # type: ignore

    def detect_documents():
        """Detect open Office/WPS documents. Returns same format as office_detect."""
        if _detect_all is None:
            return {"error": "com_resolver not available"}
        return _detect_all()

    # ── Workspace-aware save path resolution ──
    # Resolves relative paths to HANDY_WORKSPACE (set by Tauri bridge).
    # Absolute paths are returned as-is.
    import os as _os
    _workspace = _os.environ.get("HANDY_WORKSPACE", _os.getcwd())

    def _resolve_save_path(save_path):
        """Resolve save_path to absolute: relative paths → workspace, absolute → as-is."""
        if save_path is None:
            return None
        p = str(save_path)
        # Absolute path (Windows drive letter or Unix root)
        if _os.path.isabs(p):
            return p
        # Relative → workspace
        _os.makedirs(_workspace, exist_ok=True)
        resolved = _os.path.join(_workspace, p)
        print(f"[doc_code_exec] Resolved relative save path '{p}' → '{resolved}'", file=sys.stderr)
        return resolved

    # Generator convenience functions (replaces generate_doc tool)
    def generate_excel(title, sheets, save_path=None, author=None):
        """Generate .xlsx file. Relative save_path → workspace, absolute → as-is. Returns bytes if no save_path."""
        gen = _get_excel_gen()
        data = gen.generate(title=title, sheets=sheets, author=author)
        resolved = _resolve_save_path(save_path)
        if resolved:
            with open(resolved, "wb") as f:
                f.write(data)
            return {"saved": True, "path": resolved, "size": len(data)}
        import base64
        return {"saved": False, "data": base64.b64encode(data).decode(), "size": len(data)}

    def generate_word(title, content, save_path=None, subtitle=None, author=None):
        """Generate .docx file. Relative save_path → workspace, absolute → as-is. Returns bytes if no save_path."""
        gen = _get_word_gen()
        data = gen.generate(title=title, content=content, subtitle=subtitle, author=author)
        resolved = _resolve_save_path(save_path)
        if resolved:
            with open(resolved, "wb") as f:
                f.write(data)
            return {"saved": True, "path": resolved, "size": len(data)}
        import base64
        return {"saved": False, "data": base64.b64encode(data).decode(), "size": len(data)}

    def generate_ppt(title, slides=None, markdown=None, save_path=None, author=None):
        """Generate .pptx file. Relative save_path → workspace, absolute → as-is. Returns bytes if no save_path."""
        gen = _get_ppt_gen()
        if markdown:
            data = gen.generate_from_markdown(title=title, markdown=markdown, author=author)
        else:
            data = gen.generate(title=title, slides=slides or [], author=author)
        resolved = _resolve_save_path(save_path)
        if resolved:
            with open(resolved, "wb") as f:
                f.write(data)
            return {"saved": True, "path": resolved, "size": len(data)}
        import base64
        return {"saved": False, "data": base64.b64encode(data).decode(), "size": len(data)}

    sandbox_globals = {
        "__builtins__": __builtins__,
        # openpyxl
        "Workbook": _OWB,
        "load_workbook": _load_wb,
        "Font": _Font,
        "Alignment": _Align,
        "PatternFill": _Fill,
        "Border": _Border,
        "Side": _Side,
        # python-docx
        "Document": _DocxDoc,
        # python-pptx
        "Presentation": _PptxPres,
        # COM classes
        "ExcelCOM": ExcelCOM,
        "WordCOM": WordCOM,
        "PptCOM": PptCOM,
        # Generator classes
        "ExcelGenerator": ExcelGenerator,
        "WordGenerator": WordGenerator,
        "PptGenerator": PptGenerator,
        # Convenience functions
        "get_excel_app": get_excel_app,
        "get_word_app": get_word_app,
        "get_ppt_app": get_ppt_app,
        "read_range": read_range,
        "save_workbook": save_workbook,
        "save_document": save_document,
        "save_presentation": save_presentation,
        # Detection (replaces office_detect)
        "detect_documents": detect_documents,
        # Generators (replaces generate_doc)
        "generate_excel": generate_excel,
        "generate_word": generate_word,
        "generate_ppt": generate_ppt,
    }

    result_value, error_str = None, None
    start = time.time()

    # ── Execute with timeout ──
    exec_error = None

    def _run():
        nonlocal result_value, exec_error
        try:
            compiled = compile(code, "<doc_sandbox>", "exec", flags=0)
            exec(compiled, sandbox_globals)
            result_value = sandbox_globals.get("result", None)
        except Exception:
            exec_error = traceback.format_exc()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    thread.join(timeout=timeout_sec)

    if thread.is_alive():
        error_str = f"Execution timed out after {timeout_sec}s"
    elif exec_error:
        error_str = exec_error

    duration_ms = int((time.time() - start) * 1000)
    output = captured_stdout.getvalue()
    err_output = captured_stderr.getvalue()

    # Restore stdout/stderr
    sys.stdout, sys.stderr = old_stdout, old_stderr

    # Sanitize result for JSON serialization (datetime, bytes, etc.)
    def _sanitize(obj: Any, _depth: int = 0) -> Any:
        if _depth > 20:
            return str(obj)
        if obj is None or isinstance(obj, (bool, int, float, str)):
            return obj
        if isinstance(obj, bytes):
            return f"<bytes {len(obj)}>"
        if isinstance(obj, datetime):
            return obj.isoformat()
        if isinstance(obj, dict):
            return {str(k): _sanitize(v, _depth + 1) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_sanitize(v, _depth + 1) for v in obj]
        # openpyxl / python-docx / COM objects — convert to string
        try:
            json.dumps(obj)
            return obj
        except (TypeError, ValueError):
            return str(obj)

    safe_result = _sanitize(result_value)

    # Release COM connections so files aren't locked between tasks
    try:
        from engine.office.com_resolver import clear_cache as _clear_com_cache
        _clear_com_cache()
    except Exception:
        pass

    return {
        "success": error_str is None,
        "output": output[:100000],  # truncate large output
        "error": error_str or err_output or "",
        "result": safe_result,
        "duration_ms": duration_ms,
        "truncated": len(output) > 100000,
    }
