"""Handy Automation Engine — Python sidecar.

Reads JSON Line requests from stdin, dispatches to tool handlers,
writes JSON Line responses to stdout. Runs until stdin closes.

Protocol:
  Request:  {"id":"...","tool":"...","params":{...}}
  Response: {"id":"...","ok":true,"data":{...}}
            {"id":"...","ok":false,"error":"..."}
"""

from __future__ import annotations

import sys
import traceback
from typing import Any, Callable

from protocol import read_request, write_response, ok, fail
from engine.desktop_uia import UIAEngine
from engine.browser import BrowserEngine
from engine.screenshot import ScreenshotEngine
from engine.ocr import OCREngine
from engine.office import WordGenerator, ExcelGenerator, PptGenerator
from engine.web_search import WebSearchEngine
# COM modules are optional — auto-install pywin32 if missing
try:
    from engine.office import WordCOM, ExcelCOM, PptCOM
except ImportError:
    WordCOM = None  # type: ignore[assignment,misc]
    ExcelCOM = None  # type: ignore[assignment,misc]
    PptCOM = None  # type: ignore[assignment,misc]
from engine.global_listener import get_listener

# ── Tool registry ──
# Each handler receives (params: dict) -> dict (the data field of the response)

_uia: UIAEngine | None = None
_browser: BrowserEngine | None = None
_screenshot: ScreenshotEngine | None = None
_ocr: OCREngine | None = None
_word_gen: WordGenerator | None = None
_excel_gen: ExcelGenerator | None = None
_ppt_gen: PptGenerator | None = None
_word_com = None  # WordCOM instance (lazy)
_excel_com = None  # ExcelCOM instance (lazy)
_ppt_com = None  # PptCOM instance (lazy)
_web_search: WebSearchEngine | None = None


def _get_uia() -> UIAEngine:
    global _uia
    if _uia is None:
        _uia = UIAEngine()
    return _uia


def _get_browser() -> BrowserEngine:
    global _browser
    if _browser is None:
        _browser = BrowserEngine()
        # Register with extension WebSocket server
        try:
            from engine.extension_ws import set_browser_engine
            set_browser_engine(_browser)
        except ImportError:
            pass
    return _browser


def _get_extension_connected() -> bool:
    """Check if Chrome extension is connected via WebSocket."""
    try:
        from engine.extension_ws import is_extension_connected
        return is_extension_connected()
    except ImportError:
        return False


def _get_browser_status() -> dict:
    """Get browser connection status including URL if available."""
    try:
        from engine.extension_ws import get_browser_status
        return get_browser_status()
    except ImportError:
        return {"connected": False}


def _get_screenshot() -> ScreenshotEngine:
    global _screenshot
    if _screenshot is None:
        _screenshot = ScreenshotEngine()
    return _screenshot


def _get_ocr() -> OCREngine:
    global _ocr
    if _ocr is None:
        _ocr = OCREngine()
    return _ocr


def _get_word_gen() -> WordGenerator:
    global _word_gen
    if _word_gen is None:
        _word_gen = WordGenerator()
    return _word_gen


def _get_excel_gen() -> ExcelGenerator:
    global _excel_gen
    if _excel_gen is None:
        _excel_gen = ExcelGenerator()
    return _excel_gen


def _get_ppt_gen() -> PptGenerator:
    global _ppt_gen
    if _ppt_gen is None:
        _ppt_gen = PptGenerator()
    return _ppt_gen


def _get_word_com():
    global _word_com
    if WordCOM is None:
        raise RuntimeError("WordCOM unavailable — pywin32 not installed or import failed")
    if _word_com is None:
        _word_com = WordCOM()
    return _word_com


def _get_excel_com():
    global _excel_com
    if ExcelCOM is None:
        raise RuntimeError("ExcelCOM unavailable — pywin32 not installed or import failed")
    if _excel_com is None:
        _excel_com = ExcelCOM()
    return _excel_com


def _get_ppt_com():
    global _ppt_com
    if PptCOM is None:
        raise RuntimeError("PptCOM unavailable — pywin32 not installed or import failed")
    if _ppt_com is None:
        _ppt_com = PptCOM()
    return _ppt_com


def _get_web_search() -> WebSearchEngine:
    global _web_search
    if _web_search is None:
        _web_search = WebSearchEngine()
    return _web_search


def _handle_uia_get_interactive(params: dict) -> dict[str, Any]:
    hwnd = params.get("window_hwnd")
    return _get_uia().get_interactive_nodes(
        window_hwnd=hwnd,
        roles=params.get("roles"),
        name_keyword=params.get("name_keyword"),
        onscreen_only=params.get("onscreen_only", False),
        limit=params.get("limit"),
    )


def _handle_uia_click(params: dict) -> dict[str, Any]:
    return _get_uia().click(
        role=params["role"],
        name=params.get("name"),
        window_hwnd=params.get("window_hwnd"),
    )


def _handle_uia_type(params: dict) -> dict[str, Any]:
    return _get_uia().type_text(
        text=params["text"],
        role=params.get("role"),
        name=params.get("name"),
        window_hwnd=params.get("window_hwnd"),
    )


def _handle_uia_find(params: dict) -> dict[str, Any]:
    return _get_uia().find_element(
        role=params["role"],
        name=params.get("name"),
        window_hwnd=params.get("window_hwnd"),
    )


def _handle_uia_get_property(params: dict) -> dict[str, Any]:
    return _get_uia().get_property(
        role=params["role"],
        name=params.get("name"),
        prop=params["property"],
        window_hwnd=params.get("window_hwnd"),
    )


def _handle_uia_fingerprint(params: dict) -> dict[str, Any]:
    return _get_uia().fingerprint(window_hwnd=params.get("window_hwnd"))


def _handle_uia_find_at_point(params: dict) -> dict[str, Any]:
    return _get_uia().find_at_point(
        x=params["x"],
        y=params["y"],
        hwnd=params.get("hwnd"),
    )


# ── Browser (Playwright) handlers ──

def _handle_web_launch(params: dict) -> dict[str, Any]:
    return _get_browser().launch(
        headless=params.get("headless", True),
        channel=params.get("channel", ""),
        connect_existing=params.get("connect_existing", True),
    )


def _handle_web_connect_cdp(params: dict) -> dict[str, Any]:
    return _get_browser().connect_cdp(
        cdp_url=params.get("cdp_url", "http://localhost:9222"),
    )


def _handle_web_navigate(params: dict) -> dict[str, Any]:
    """Navigate to URL. Uses Playwright if available, otherwise uses extension."""
    browser_engine = _get_browser()

    # Try Playwright first
    if browser_engine._page:
        return browser_engine.navigate(
            url=params.get("url", ""),
            action=params.get("action", "goto"),
        )

    # Fall back to extension
    try:
        from engine.extension_ws import send_to_extension_sync
        url = params.get("url", "")
        if not url:
            return {"navigated": False, "error": "URL is required"}
        js_code = f"window.location.href = '{url}'"
        result = send_to_extension_sync("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return {"navigated": True, "url": url}
        return {"navigated": False, "error": result.get("error", "Extension execution failed")}
    except Exception as e:
        return {"navigated": False, "error": str(e)}


def _handle_web_get_interactive(_params: dict) -> dict[str, Any]:
    """Get interactive elements from the page.

    Uses Playwright if available, otherwise uses extension.
    """
    browser_engine = _get_browser()

    # Try Playwright first
    if browser_engine._page:
        return browser_engine.get_interactive_nodes()

    # Fall back to extension — use predefined ext_get_interactive (no eval, CSP-safe)
    try:
        from engine.extension_ws import send_to_extension_sync
        result = send_to_extension_sync("ext_get_interactive", {})
        if result.get("ok"):
            data = result["data"]["results"][0] if result["data"]["results"] else {}
            return {
                "url": data.get("url", ""),
                "title": data.get("title", ""),
                "nodes": data.get("nodes", []),
                "count": len(data.get("nodes", [])),
            }
        return {"url": "", "title": "", "nodes": [], "count": 0, "error": result.get("error", "Extension execution failed")}
    except Exception as e:
        return {"url": "", "title": "", "nodes": [], "count": 0, "error": str(e)}


def _handle_ext_get_recorded_events(_params: dict) -> dict[str, Any]:
    """Get user interaction events captured by the Chrome extension content script.

    Drains the event buffer (events are removed after being returned).
    """
    try:
        from engine.extension_ws import send_to_extension_sync
        result = send_to_extension_sync("ext_get_recorded_events", {})
        if result.get("ok"):
            return {"events": result["data"].get("events", []), "ok": True}
        return {"events": [], "ok": False, "error": result.get("error", "Extension not connected")}
    except Exception as e:
        return {"events": [], "ok": False, "error": str(e)}


def _handle_ext_set_capture(params: dict) -> dict[str, Any]:
    """Enable or disable event capture in the extension content scripts."""
    enabled = params.get("enabled", False)
    try:
        from engine.extension_ws import send_to_extension_sync
        result = send_to_extension_sync("ext_set_capture", {"enabled": enabled})
        if result.get("ok"):
            return {"ok": True, "enabled": result["data"].get("enabled", enabled)}
        return {"ok": False, "error": result.get("error", "Extension not connected")}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def _handle_web_click_selector(params: dict) -> dict[str, Any]:
    """Click element by selector. Uses Playwright if available, otherwise uses extension."""
    browser_engine = _get_browser()
    if browser_engine._page:
        return browser_engine.click_selector(selector=params["selector"])

    # Fall back to extension
    try:
        from engine.extension_ws import send_to_extension_sync
        selector = params["selector"]
        js_code = f"document.querySelector('{selector}').click()"
        result = send_to_extension_sync("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return {"clicked": True, "selector": selector}
        return {"clicked": False, "error": result.get("error", "Extension execution failed")}
    except Exception as e:
        return {"clicked": False, "error": str(e)}


def _handle_web_click_role(params: dict) -> dict[str, Any]:
    """Click element by role. Uses Playwright if available, otherwise uses extension."""
    browser_engine = _get_browser()
    if browser_engine._page:
        return browser_engine.click_role(
            role=params["role"],
            name=params.get("name"),
        )

    # Fall back to extension
    try:
        from engine.extension_ws import send_to_extension_sync
        role = params["role"]
        name = params.get("name", "")
        if name:
            js_code = f"document.querySelector('[role=\"{role}\"][aria-label=\"{name}\"]').click()"
        else:
            js_code = f"document.querySelector('[role=\"{role}\"]').click()"
        result = send_to_extension_sync("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return {"clicked": True, "role": role, "name": name}
        return {"clicked": False, "error": result.get("error", "Extension execution failed")}
    except Exception as e:
        return {"clicked": False, "error": str(e)}


def _handle_web_fill(params: dict) -> dict[str, Any]:
    """Fill input field. Uses Playwright if available, otherwise uses extension."""
    browser_engine = _get_browser()
    if browser_engine._page:
        return browser_engine.fill(selector=params["selector"], text=params["text"])

    # Fall back to extension
    try:
        from engine.extension_ws import send_to_extension_sync
        selector = params["selector"]
        text = params["text"].replace("'", "\\'").replace("\n", "\\n")
        js_code = f"document.querySelector('{selector}').value = '{text}'"
        result = send_to_extension_sync("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return {"filled": True, "selector": selector}
        return {"filled": False, "error": result.get("error", "Extension execution failed")}
    except Exception as e:
        return {"filled": False, "error": str(e)}


def _handle_web_type(params: dict) -> dict[str, Any]:
    """Type text. Uses Playwright if available, otherwise uses extension."""
    browser_engine = _get_browser()
    if browser_engine._page:
        return browser_engine.type_text(text=params["text"])

    # Fall back to extension - type into focused element
    try:
        from engine.extension_ws import send_to_extension_sync
        text = params["text"].replace("'", "\\'").replace("\n", "\\n")
        js_code = f"document.activeElement.value += '{text}'"
        result = send_to_extension_sync("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return {"typed": True}
        return {"typed": False, "error": result.get("error", "Extension execution failed")}
    except Exception as e:
        return {"typed": False, "error": str(e)}


def _handle_web_scroll(params: dict) -> dict[str, Any]:
    """Scroll page. Uses Playwright if available, otherwise uses extension."""
    browser_engine = _get_browser()
    if browser_engine._page:
        return browser_engine.scroll(delta_y=params.get("delta_y", 300))

    # Fall back to extension
    try:
        from engine.extension_ws import send_to_extension_sync
        delta_y = params.get("delta_y", 300)
        js_code = f"window.scrollBy(0, {delta_y})"
        result = send_to_extension_sync("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return {"scrolled": True, "delta_y": delta_y}
        return {"scrolled": False, "error": result.get("error", "Extension execution failed")}
    except Exception as e:
        return {"scrolled": False, "error": str(e)}


def _handle_web_close(_params: dict) -> dict[str, Any]:
    global _browser
    if _browser is None:
        return {"closed": True, "error": "No browser instance"}
    result = _browser.close()
    # Clear the global reference so a fresh BrowserEngine is created next time
    _browser = None
    return result


def _handle_web_launch_browser(params: dict) -> dict[str, Any]:
    """Launch a browser with remote debugging port and connect to it.

    Strategy:
    1. Try connecting to an existing browser debug port (user may already have Edge open)
    2. If not found, launch with real profile (user's cookies/logins)
    3. If real profile is locked (browser already running), fall back to temp profile
    """
    browser = params.get("browser", "msedge")
    port = params.get("port", 9222)
    engine = _get_browser()

    # Step 1: Try connecting to an existing browser debug port
    existing = engine._try_connect_existing()
    if existing and existing.get("connected"):
        existing["method"] = "connect_existing"
        return existing

    # Step 2: Launch with real profile (has user's cookies, logins, bookmarks)
    result = engine._launch_with_debug_port(browser, port, use_real_profile=True)
    if result.get("launched"):
        cdp_url = f"http://127.0.0.1:{port}"
        connect_result = engine.connect_cdp(cdp_url)
        if connect_result.get("connected"):
            connect_result["method"] = "launch_with_debug"
            return connect_result

    # Step 3: Real profile failed (locked by running browser), try temp profile
    result = engine._launch_with_debug_port(browser, port, use_real_profile=False)
    if result.get("launched"):
        cdp_url = f"http://127.0.0.1:{port}"
        connect_result = engine.connect_cdp(cdp_url)
        if connect_result.get("connected"):
            connect_result["method"] = "launch_with_debug"
            return connect_result
    return result


def _handle_web_start_recording(_params: dict) -> dict[str, Any]:
    result = _get_browser().start_recording()
    # 同时启用扩展的事件捕获
    try:
        from engine.extension_ws import send_to_extension_sync, is_extension_connected
        if is_extension_connected():
            send_to_extension_sync("ext_set_capture", {"enabled": True})
    except Exception:
        pass
    return result


def _handle_web_stop_recording(_params: dict) -> dict[str, Any]:
    result = _get_browser().stop_recording()
    # 同时禁用扩展的事件捕获
    try:
        from engine.extension_ws import send_to_extension_sync, is_extension_connected
        if is_extension_connected():
            send_to_extension_sync("ext_set_capture", {"enabled": False})
    except Exception:
        pass
    return result


def _handle_web_get_recorded_events(_params: dict) -> dict[str, Any]:
    return _get_browser().get_recorded_events()


# ── Screenshot (mss) handlers ──

def _handle_screenshot_full(_params: dict) -> dict[str, Any]:
    return _get_screenshot().full()


def _handle_screenshot_region(params: dict) -> dict[str, Any]:
    return _get_screenshot().region(
        left=params["left"],
        top=params["top"],
        width=params["width"],
        height=params["height"],
    )


def _handle_screenshot_monitors(_params: dict) -> dict[str, Any]:
    return _get_screenshot().all_monitors()


# ── OCR (PaddleOCR) handlers ──

def _handle_ocr_recognize(params: dict) -> dict[str, Any]:
    return _get_ocr().recognize(
        image_path=params.get("image_path", ""),
        image_base64=params.get("image_base64", ""),
    )


# ── Global input listener (pynput) handlers ──

def _handle_global_listener_start(params: dict) -> dict[str, Any]:
    listener = get_listener()
    listener.start(parent_pid=params.get("parent_pid", 0))
    return {"running": listener.is_running}


def _handle_global_listener_stop(_params: dict) -> dict[str, Any]:
    listener = get_listener()
    listener.stop()
    return {"running": listener.is_running}


def _handle_global_listener_poll(params: dict) -> dict[str, Any]:
    listener = get_listener()
    max_events = params.get("max_events", 100)
    events = listener.poll(max_events=max_events)
    return {"events": events, "count": len(events)}


def _handle_event_collector_poll(params: dict) -> dict[str, Any]:
    from engine.event_collector import get_collector
    collector = get_collector()
    max_events = params.get("max_events", 50)
    events = collector.poll(max_events=max_events)
    return {"events": events, "count": len(events)}


def _handle_event_collector_start(_params: dict) -> dict[str, Any]:
    from engine.event_collector import get_collector
    collector = get_collector()
    collector.start_recording()
    return {"recording": collector.is_recording}


def _handle_event_collector_stop(_params: dict) -> dict[str, Any]:
    from engine.event_collector import get_collector
    collector = get_collector()
    collector.stop_recording()
    return {"recording": collector.is_recording}


# ── Office document generators ──

def _handle_word_generate(params: dict) -> dict[str, Any]:
    import base64
    gen = _get_word_gen()
    docx_bytes = gen.generate(
        title=params["title"],
        content=params.get("content", ""),
        subtitle=params.get("subtitle"),
        author=params.get("author"),
    )
    save_path = params.get("save_path")
    if save_path:
        with open(save_path, "wb") as f:
            f.write(docx_bytes)
        return {"saved": True, "path": save_path, "size": len(docx_bytes)}
    return {"saved": False, "data": base64.b64encode(docx_bytes).decode("utf-8"), "size": len(docx_bytes)}


def _handle_excel_generate(params: dict) -> dict[str, Any]:
    import base64
    gen = _get_excel_gen()
    xlsx_bytes = gen.generate(
        title=params["title"],
        sheets=params.get("sheets", []),
        author=params.get("author"),
    )
    save_path = params.get("save_path")
    if save_path:
        with open(save_path, "wb") as f:
            f.write(xlsx_bytes)
        return {"saved": True, "path": save_path, "size": len(xlsx_bytes)}
    return {"saved": False, "data": base64.b64encode(xlsx_bytes).decode("utf-8"), "size": len(xlsx_bytes)}


def _handle_ppt_generate(params: dict) -> dict[str, Any]:
    import base64
    gen = _get_ppt_gen()

    # Support both structured slides and markdown content
    if "markdown" in params:
        pptx_bytes = gen.generate_from_markdown(
            title=params["title"],
            markdown=params["markdown"],
            author=params.get("author"),
        )
    else:
        pptx_bytes = gen.generate(
            title=params["title"],
            slides=params.get("slides", []),
            author=params.get("author"),
        )

    save_path = params.get("save_path")
    if save_path:
        with open(save_path, "wb") as f:
            f.write(pptx_bytes)
        return {"saved": True, "path": save_path, "size": len(pptx_bytes)}
    return {"saved": False, "data": base64.b64encode(pptx_bytes).decode("utf-8"), "size": len(pptx_bytes)}


# ── Office COM automation handlers ──

def _handle_office_detect(_params: dict) -> dict[str, Any]:
    """Detect active Word/Excel/PPT documents using the resolver."""
    from engine.office.com_resolver import detect_all
    return detect_all()


def _handle_word_com_read(params: dict) -> dict[str, Any]:
    return _get_word_com().read_content(
        paragraph_start=params.get("paragraph_start"),
        paragraph_end=params.get("paragraph_end"),
    )


def _handle_word_com_edit(params: dict) -> dict[str, Any]:
    com = _get_word_com()
    op = params.get("operation", "replace")
    if op == "open":
        return com.open(params["file_path"])
    elif op == "sync":
        return com.sync()
    elif op == "save":
        return com.save()
    elif op == "replace":
        return com.replace_text(params["find"], params["replace"])
    elif op == "set_paragraph":
        return com.set_paragraph(params["paragraph_index"], params["text"])
    elif op == "insert":
        return com.insert_text(params["after_paragraph"], params["text"])
    elif op == "insert_heading":
        return com.insert_heading(
            params["after_paragraph"], params["text"], params.get("level", 1)
        )
    elif op == "delete":
        return com.delete_paragraph(params["paragraph_index"])
    elif op == "format":
        return com.apply_format(
            params["paragraph_index"],
            bold=params.get("bold"),
            italic=params.get("italic"),
            font_size=params.get("font_size"),
        )
    elif op == "get_selection":
        return com.get_selection()
    else:
        return {"success": False, "message": f"Unknown operation: {op}"}


def _handle_excel_com_read(params: dict) -> dict[str, Any]:
    com = _get_excel_com()
    file_path = params.get("file_path")
    if params.get("get_selection"):
        return com.get_selection()
    range_addr = params.get("range")
    if range_addr:
        return com.read_range(
            range_addr=range_addr,
            sheet=params.get("sheet"),
            file_path=file_path,
        )
    # 没传 range 时默认返回 sheet_info
    return com.get_sheet_info(params.get("sheet"), file_path=file_path)


def _handle_excel_com_edit(params: dict) -> dict[str, Any]:
    com = _get_excel_com()
    op = params.get("operation", "write")
    if op == "open":
        return com.open(params["file_path"])
    elif op == "sync":
        return com.sync()
    elif op == "save":
        return com.save()
    elif op == "write":
        return com.write_range(params["range"], params["values"], params.get("sheet"))
    elif op == "formula":
        return com.set_formula(params["cell"], params["formula"], params.get("sheet"))
    elif op == "auto_fill":
        return com.auto_fill_column(
            params["column"], params["formula_template"],
            params["start_row"], params["end_row"], params.get("sheet"),
        )
    elif op == "set_value":
        return com.set_value(params["cell"], params["value"], params.get("sheet"))
    elif op == "format":
        return com.format_column(
            params["column"],
            number_format=params.get("number_format"),
            bold_header=params.get("bold_header", False),
            sheet=params.get("sheet"),
        )
    elif op == "insert_rows":
        return com.insert_rows(params["after_row"], params.get("count", 1), params.get("sheet"))
    elif op == "insert_columns":
        return com.insert_columns(params["after_col"], params.get("count", 1), params.get("sheet"))
    else:
        return {"success": False, "message": f"Unknown operation: {op}"}


def _handle_ppt_com_read(params: dict) -> dict[str, Any]:
    com = _get_ppt_com()
    if params.get("slide_info"):
        return com.read_slide(params["slide_index"])
    if params.get("find_text"):
        return com.find_text_shapes(params.get("slide_index"))
    return com.read_content(
        slide_start=params.get("slide_start"),
        slide_end=params.get("slide_end"),
    )


def _handle_ppt_com_edit(params: dict) -> dict[str, Any]:
    com = _get_ppt_com()
    op = params.get("operation", "set_text")
    if op == "open":
        return com.open(params["file_path"])
    elif op == "sync":
        return com.sync()
    elif op == "save":
        return com.save()
    elif op == "set_text":
        return com.set_slide_text(params["slide_index"], params["shape_name"], params["text"])
    elif op == "add_slide":
        return com.add_slide(
            params.get("layout_index", 1),
            params.get("title"),
            params.get("content"),
            params.get("after_slide"),
        )
    elif op == "delete_slide":
        return com.delete_slide(params["slide_index"])
    elif op == "reorder":
        return com.reorder_slides(params["new_order"])
    else:
        return {"success": False, "message": f"Unknown operation: {op}"}


def _handle_exec_python(params: dict) -> dict:
    """Execute arbitrary Python code in a restricted sandbox.

    Allowed modules (SAFE_MODULES):
      json, math, datetime, re, collections, itertools, random, statistics,
      uuid, base64, hashlib, textwrap, string, typing, enum, functools,
      operator, bisect, decimal, fractions, copy.

    The value of a variable named ``result`` in the executed scope is
    returned as the ``result`` field.
    """
    code = params.get("code", "")
    timeout_sec = params.get("timeout_sec", 30)
    input_vars = params.get("params", {})

    import io
    import sys
    import traceback
    import time

    old_stdout, old_stderr = sys.stdout, sys.stderr
    captured_stdout, captured_stderr = io.StringIO(), io.StringIO()
    sys.stdout, sys.stderr = captured_stdout, captured_stderr

    SAFE_MODULES = {
        "json", "math", "datetime", "re", "collections", "itertools",
        "random", "statistics", "uuid", "base64", "hashlib", "textwrap",
        "string", "typing", "enum", "functools", "operator", "bisect",
        "decimal", "fractions", "copy",
    }

    def safe_import(name, *args):
        if name not in SAFE_MODULES:
            raise ImportError(f"Module '{name}' is not allowed")
        return __import__(name, *args)

    safe_globals = {
        "__builtins__": {
            "print": print,
            "len": len,
            "range": range,
            "int": int,
            "float": float,
            "str": str,
            "bool": bool,
            "list": list,
            "dict": dict,
            "tuple": tuple,
            "set": set,
            "sorted": sorted,
            "reversed": reversed,
            "enumerate": enumerate,
            "zip": zip,
            "map": map,
            "filter": filter,
            "any": any,
            "all": all,
            "min": min,
            "max": max,
            "sum": sum,
            "abs": abs,
            "round": round,
            "isinstance": isinstance,
            "type": type,
            "hasattr": hasattr,
            "getattr": getattr,
            "setattr": setattr,
            "ValueError": ValueError,
            "TypeError": TypeError,
            "KeyError": KeyError,
            "IndexError": IndexError,
            "Exception": Exception,
            "StopIteration": StopIteration,
            "__import__": safe_import,
        }
    }
    safe_globals.update(input_vars)

    result_value, error_str = None, None
    start = time.time()

    try:
        compiled = compile(code, "<sandbox>", "exec", flags=0)
        exec(compiled, safe_globals)
        result_value = safe_globals.get("result", None)
    except Exception:
        error_str = traceback.format_exc()
    finally:
        sys.stdout, sys.stderr = old_stdout, old_stderr

    duration_ms = int((time.time() - start) * 1000)
    output = captured_stdout.getvalue()
    err_output = captured_stderr.getvalue()

    return {
        "success": error_str is None,
        "output": output,
        "error": error_str or err_output or "",
        "result": result_value,
        "duration_ms": duration_ms,
        "truncated": len(output) > 100000,
    }


def _handle_web_code_exec(params: dict) -> dict:
    """Execute Python code with Playwright browser context injected.

    Pre-injected variables:
      - page: Playwright Page object (current active page)
      - browser: Playwright Browser object
      - wait_for: convenience wrapper for page.wait_for_selector
      - screenshot_b64(): take screenshot and return base64 string
      - navigate/click/fill/scroll/get_interactive/get_content/evaluate: helper functions

    The value of a variable named ``result`` is returned as the ``result`` field.
    """
    code = params.get("code", "")
    timeout_sec = params.get("timeout_sec", 60)

    import io
    import sys
    import time
    import traceback
    import base64

    browser_engine = _get_browser()

    # Check if we have Playwright page OR extension connection
    has_playwright = browser_engine._page is not None
    has_extension = False
    try:
        from engine.extension_ws import is_extension_connected
        has_extension = is_extension_connected()
    except ImportError:
        pass

    if not has_playwright and not has_extension:
        return {
            "success": False,
            "output": "",
            "error": "No browser available. Connect browser extension or call web_launch first.",
            "result": None,
            "duration_ms": 0,
            "truncated": False,
        }

    page = browser_engine._page if has_playwright else None
    browser = browser_engine._browser if has_playwright else None

    # ── Dangerous operation guard ──
    _DANGEROUS_PATTERNS = [
        "os._exit", "os.system", "os.exec", "os.spawn",
        "os.kill", "os.remove", "os.unlink", "os.rmdir",
        "shutil.rmtree", "shutil.move",
        "subprocess.", "ctypes.",
        "importlib.",
    ]
    for pat in _DANGEROUS_PATTERNS:
        if pat in code:
            return {
                "success": False,
                "output": "",
                "error": f"Blocked dangerous operation: '{pat}' is not allowed in web_code_exec.",
                "result": None,
                "duration_ms": 0,
                "truncated": False,
            }

    # ── Capture stdout/stderr ──
    old_stdout, old_stderr = sys.stdout, sys.stderr
    captured_stdout, captured_stderr = io.StringIO(), io.StringIO()
    sys.stdout, sys.stderr = captured_stdout, captured_stderr

    # ── Extension helper (when Playwright is not available) ──
    def _ext_exec(js_code):
        """Execute JavaScript via extension and return result."""
        from engine.extension_ws import _handle_command
        result = _handle_command("ext_execute_script", {"code": js_code})
        if result.get("ok"):
            return result["data"]["results"][0] if result["data"]["results"] else None
        raise Exception(result.get("error", "Extension execution failed"))

    # ── Helper functions (work with both Playwright and extension) ──
    def screenshot_b64():
        """Take a screenshot and return as base64 string."""
        if has_playwright:
            img_bytes = page.screenshot()
            return base64.b64encode(img_bytes).decode("ascii")
        raise Exception("Screenshot requires Playwright browser. Call web_launch first.")

    def wait_for(selector, timeout=10000):
        """Wait for selector to appear, return the element."""
        if has_playwright:
            return page.wait_for_selector(selector, timeout=timeout)
        # Extension mode: poll with JavaScript
        import time as _time
        deadline = _time.time() + timeout / 1000
        while _time.time() < deadline:
            exists = _ext_exec(f"!!document.querySelector('{selector}')")
            if exists:
                return True
            _time.sleep(0.5)
        raise TimeoutError(f"Selector '{selector}' not found within {timeout}ms")

    def navigate(url, wait_until="load"):
        """Navigate to URL."""
        if has_playwright:
            page.goto(url, wait_until=wait_until)
            return {"url": page.url, "title": page.title()}
        _ext_exec(f"window.location.href = '{url}'")
        return {"url": url, "navigated": True}

    def click(selector):
        """Click element by selector."""
        if has_playwright:
            page.click(selector)
            return {"clicked": selector}
        _ext_exec(f"document.querySelector('{selector}').click()")
        return {"clicked": selector}

    def fill(selector, value):
        """Fill input field."""
        if has_playwright:
            page.fill(selector, value)
            return {"filled": selector}
        escaped = value.replace("'", "\\'").replace("\n", "\\n")
        _ext_exec(f"document.querySelector('{selector}').value = '{escaped}'")
        return {"filled": selector}

    def scroll(direction="down", amount=500):
        """Scroll page. direction: 'up', 'down', 'left', 'right'."""
        js = {
            "down": f"window.scrollBy(0, {amount})",
            "up": f"window.scrollBy(0, -{amount})",
            "right": f"window.scrollBy({amount}, 0)",
            "left": f"window.scrollBy(-{amount}, 0)",
        }.get(direction, f"window.scrollBy(0, {amount})")
        if has_playwright:
            page.evaluate(js)
        else:
            _ext_exec(js)
        return {"scrolled": direction, "amount": amount}

    def get_interactive():
        """Get all interactive elements on the page with selectors."""
        js = """() => {
            const elements = document.querySelectorAll('a, button, input, select, textarea, [role="button"], [onclick]');
            return Array.from(elements).map(el => ({
                tag: el.tagName.toLowerCase(),
                text: el.textContent?.trim()?.substring(0, 100) || '',
                id: el.id || '',
                name: el.name || '',
                type: el.type || '',
                href: el.href || '',
                placeholder: el.placeholder || '',
                value: el.value || '',
                selector: el.id ? `#${el.id}` : el.name ? `[name="${el.name}"]` : null
            })).filter(e => e.text || e.id || e.name || e.placeholder);
        }"""
        if has_playwright:
            return page.evaluate(js)
        return _ext_exec(js)

    def get_content(text_only=True, max_length=50000):
        """Get page content. Set text_only=False for HTML."""
        if has_playwright:
            if text_only:
                content = page.evaluate("document.body.innerText")
            else:
                content = page.content()
        else:
            js = "document.body.innerText" if text_only else "document.documentElement.outerHTML"
            content = _ext_exec(js)
        return content[:max_length] if len(content) > max_length else content

    def evaluate(js_code):
        """Execute JavaScript in the page context and return result."""
        if has_playwright:
            return page.evaluate(js_code)
        return _ext_exec(js_code)

    def close_browser():
        """Close the browser."""
        if has_playwright:
            browser.close()
            return {"closed": True}
        return {"closed": False, "reason": "Extension mode - cannot close browser"}

    # ── Build sandbox globals ──
    safe_globals = {
        "__builtins__": {
            "print": print,
            "len": len,
            "range": range,
            "int": int,
            "float": float,
            "str": str,
            "bool": bool,
            "list": list,
            "dict": dict,
            "tuple": tuple,
            "set": set,
            "sorted": sorted,
            "reversed": reversed,
            "enumerate": enumerate,
            "zip": zip,
            "map": map,
            "filter": filter,
            "any": any,
            "all": all,
            "min": min,
            "max": max,
            "sum": sum,
            "abs": abs,
            "round": round,
            "isinstance": isinstance,
            "type": type,
            "hasattr": hasattr,
            "getattr": getattr,
            "setattr": setattr,
            "ValueError": ValueError,
            "TypeError": TypeError,
            "KeyError": KeyError,
            "IndexError": IndexError,
            "Exception": Exception,
            "StopIteration": StopIteration,
            "__import__": lambda name, *args: __import__(name, *args) if name in {"time", "json", "re", "base64", "math"} else (_ for _ in ()).throw(ImportError(f"Module '{name}' is not allowed")),
        },
        "page": page,
        "browser": browser,
        "wait_for": wait_for,
        "screenshot_b64": screenshot_b64,
        # Web operation helpers
        "navigate": navigate,
        "click": click,
        "fill": fill,
        "scroll": scroll,
        "get_interactive": get_interactive,
        "get_content": get_content,
        "evaluate": evaluate,
        "close_browser": close_browser,
    }

    result_value, error_str = None, None
    start = time.time()

    try:
        # Execute with timeout
        import threading

        def run_code():
            nonlocal result_value, error_str
            try:
                compiled = compile(code, "<web_sandbox>", "exec", flags=0)
                exec(compiled, safe_globals)
                result_value = safe_globals.get("result", None)
            except Exception:
                error_str = traceback.format_exc()

        thread = threading.Thread(target=run_code)
        thread.start()
        thread.join(timeout=timeout_sec)

        if thread.is_alive():
            error_str = f"Code execution timed out after {timeout_sec}s"
    except Exception:
        error_str = traceback.format_exc()
    finally:
        sys.stdout, sys.stderr = old_stdout, old_stderr

    duration_ms = int((time.time() - start) * 1000)
    output = captured_stdout.getvalue()
    err_output = captured_stderr.getvalue()

    return {
        "success": error_str is None,
        "output": output,
        "error": error_str or err_output or "",
        "result": result_value,
        "duration_ms": duration_ms,
        "truncated": len(output) > 100000,
    }


def _handle_doc_code_exec(params: dict) -> dict:
    """Execute Python code with full access to document libraries and COM instances.

    Pre-injected variables available in user code:
      - openpyxl: Workbook, load_workbook
      - docx: Document (python-docx)
      - pptx: Presentation (python-pptx)
      - ExcelCOM, WordCOM, PptCOM: COM classes
      - ExcelGenerator, WordGenerator, PptGenerator: file generators
      - get_excel_app(): get or create ExcelCOM instance (connects to user's open Excel)
      - get_word_app(): get or create WordCOM instance
      - get_ppt_app(): get or create PptCOM instance
      - read_range(addr, sheet=None, file_path=None): convenience for ExcelCOM.read_range
      - save_workbook(): convenience for ExcelCOM.save

    The value of a variable named ``result`` is returned as the ``result`` field.
    """
    code = params.get("code", "")
    timeout_sec = params.get("timeout_sec", 60)

    import io
    import sys
    import time
    import threading
    import json
    from datetime import datetime

    # ── Dangerous operation guard ──
    # Block patterns that could harm the system. This is a safety net, not a
    # hard sandbox — the LLM is trusted but we prevent accidents.
    _DANGEROUS_PATTERNS = [
        # Process / system
        "os._exit", "os.system", "os.exec", "os.spawn", "os.popen",
        "os.kill", "os.remove", "os.unlink", "os.rmdir",
        "os.rename", "os.chmod", "os.chown", "os.symlink",
        "shutil.rmtree", "shutil.move", "shutil.copy", "shutil.copytree",
        "subprocess.", "ctypes.",
        "importlib.", "pkgutil.",
        # Networking (docs don't need it)
        "socket.", "urllib.request", "urllib.error",
        "http.client", "http.server",
        "requests.get", "requests.post",
        "ftplib.", "smtplib.",
        # Code execution
        "exec(", "eval(", "compile(",
        "__builtins__",
        "globals()", "locals()",
        # Registry (Windows)
        "winreg.", "win32api.Reg",
        # Threading abuse
        "multiprocessing.",
        # Dynamic import evasion
        "__import__(",
    ]
    for pat in _DANGEROUS_PATTERNS:
        if pat in code:
            return {
                "success": False,
                "output": "",
                "error": f"Blocked dangerous operation: '{pat}' is not allowed in doc_code_exec. "
                         f"Use the predefined com_read/com_edit tools for file operations, "
                         f"or doc_code_exec for document data processing only.",
                "result": None,
                "duration_ms": 0,
                "truncated": False,
            }

    # ── Volume limit guard ──
    # Prevent denial-of-service via huge loops or memory allocation
    if len(code) > 20000:
        return {
            "success": False,
            "output": "",
            "error": "Code exceeds maximum length of 20000 characters.",
            "result": None,
            "duration_ms": 0,
            "truncated": False,
        }

    # ── Capture stdout/stderr ──
    old_stdout, old_stderr = sys.stdout, sys.stderr
    captured_stdout, captured_stderr = io.StringIO(), io.StringIO()
    sys.stdout, sys.stderr = captured_stdout, captured_stderr

    # ── Build sandbox globals with document libraries ──
    try:
        from openpyxl import Workbook as _OWB, load_workbook as _load_wb
        from openpyxl.styles import Font as _Font, Alignment as _Align, PatternFill as _Fill, Border as _Border, Side as _Side
    except ImportError:
        _OWB = None  # type: ignore
        _load_wb = None  # type: ignore
        _Font = _Align = _Fill = _Border = _Side = None

    try:
        from docx import Document as _DocxDoc
    except ImportError:
        _DocxDoc = None  # type: ignore

    try:
        from pptx import Presentation as _PptxPres
    except ImportError:
        _PptxPres = None  # type: ignore

    # Convenience: lazy COM instance getters
    _com_cache: dict = {}

    def get_excel_app():
        if "excel" not in _com_cache:
            _com_cache["excel"] = _get_excel_com()
        return _com_cache["excel"]

    def get_word_app():
        if "word" not in _com_cache:
            _com_cache["word"] = _get_word_com()
        return _com_cache["word"]

    def get_ppt_app():
        if "ppt" not in _com_cache:
            _com_cache["ppt"] = _get_ppt_com()
        return _com_cache["ppt"]

    def read_range(addr, sheet=None, file_path=None):
        return get_excel_app().read_range(addr, sheet=sheet, file_path=file_path)

    def save_workbook():
        return get_excel_app().save()

    def save_document():
        return get_word_app().save()

    def save_presentation():
        return get_ppt_app().save()

    # Document detection (replaces office_detect tool)
    try:
        from engine.office.com_resolver import detect_all as _detect_all
    except ImportError:
        _detect_all = None  # type: ignore

    def detect_documents():
        """Detect open Office/WPS documents. Returns same format as office_detect."""
        if _detect_all is None:
            return {"error": "com_resolver not available"}
        return _detect_all()

    # Generator convenience functions (replaces generate_doc tool)
    def generate_excel(title, sheets, save_path=None, author=None):
        """Generate .xlsx file. Returns bytes if no save_path."""
        gen = _get_excel_gen()
        data = gen.generate(title=title, sheets=sheets, author=author)
        if save_path:
            with open(save_path, "wb") as f:
                f.write(data)
            return {"saved": True, "path": save_path, "size": len(data)}
        import base64
        return {"saved": False, "data": base64.b64encode(data).decode(), "size": len(data)}

    def generate_word(title, content, save_path=None, subtitle=None, author=None):
        """Generate .docx file. Returns bytes if no save_path."""
        gen = _get_word_gen()
        data = gen.generate(title=title, content=content, subtitle=subtitle, author=author)
        if save_path:
            with open(save_path, "wb") as f:
                f.write(data)
            return {"saved": True, "path": save_path, "size": len(data)}
        import base64
        return {"saved": False, "data": base64.b64encode(data).decode(), "size": len(data)}

    def generate_ppt(title, slides=None, markdown=None, save_path=None, author=None):
        """Generate .pptx file. Returns bytes if no save_path."""
        gen = _get_ppt_gen()
        if markdown:
            data = gen.generate_from_markdown(title=title, markdown=markdown, author=author)
        else:
            data = gen.generate(title=title, slides=slides or [], author=author)
        if save_path:
            with open(save_path, "wb") as f:
                f.write(data)
            return {"saved": True, "path": save_path, "size": len(data)}
        import base64
        return {"saved": False, "data": base64.b64encode(data).decode(), "size": len(data)}

    sandbox_globals = {
        "__builtins__": __builtins__,
        # openpyxl
        "Workbook": _OWB,
        "load_workbook": _load_wb,
        "Font": _Font,
        "Alignment": _Align,
        "PatternFill": _Fill,
        "Border": _Border,
        "Side": _Side,
        # python-docx
        "Document": _DocxDoc,
        # python-pptx
        "Presentation": _PptxPres,
        # COM classes
        "ExcelCOM": ExcelCOM,
        "WordCOM": WordCOM,
        "PptCOM": PptCOM,
        # Generator classes
        "ExcelGenerator": ExcelGenerator,
        "WordGenerator": WordGenerator,
        "PptGenerator": PptGenerator,
        # Convenience functions
        "get_excel_app": get_excel_app,
        "get_word_app": get_word_app,
        "get_ppt_app": get_ppt_app,
        "read_range": read_range,
        "save_workbook": save_workbook,
        "save_document": save_document,
        "save_presentation": save_presentation,
        # Detection (replaces office_detect)
        "detect_documents": detect_documents,
        # Generators (replaces generate_doc)
        "generate_excel": generate_excel,
        "generate_word": generate_word,
        "generate_ppt": generate_ppt,
    }

    result_value, error_str = None, None
    start = time.time()

    # ── Execute with timeout ──
    exec_error = None

    def _run():
        nonlocal result_value, exec_error
        try:
            compiled = compile(code, "<doc_sandbox>", "exec", flags=0)
            exec(compiled, sandbox_globals)
            result_value = sandbox_globals.get("result", None)
        except Exception:
            exec_error = traceback.format_exc()

    thread = threading.Thread(target=_run, daemon=True)
    thread.start()
    thread.join(timeout=timeout_sec)

    if thread.is_alive():
        # Timeout — thread is still running
        error_str = f"Execution timed out after {timeout_sec}s"
    elif exec_error:
        error_str = exec_error

    duration_ms = int((time.time() - start) * 1000)
    output = captured_stdout.getvalue()
    err_output = captured_stderr.getvalue()

    # Restore stdout/stderr
    sys.stdout, sys.stderr = old_stdout, old_stderr

    # Sanitize result for JSON serialization (datetime, bytes, etc.)
    def _sanitize(obj: Any, _depth: int = 0) -> Any:
        if _depth > 20:
            return str(obj)
        if obj is None or isinstance(obj, (bool, int, float, str)):
            return obj
        if isinstance(obj, bytes):
            return f"<bytes {len(obj)}>"
        if isinstance(obj, datetime):
            return obj.isoformat()
        if isinstance(obj, dict):
            return {str(k): _sanitize(v, _depth + 1) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [_sanitize(v, _depth + 1) for v in obj]
        # openpyxl / python-docx / COM objects — convert to string
        try:
            json.dumps(obj)
            return obj
        except (TypeError, ValueError):
            return str(obj)

    safe_result = _sanitize(result_value)

    # Release COM connections so files aren't locked between tasks
    try:
        from engine.office.com_resolver import clear_cache as _clear_com_cache
        _clear_com_cache()
    except Exception:
        pass

    return {
        "success": error_str is None,
        "output": output[:100000],  # truncate large output
        "error": error_str or err_output or "",
        "result": safe_result,
        "duration_ms": duration_ms,
        "truncated": len(output) > 100000,
    }


# ── Web search ──

def _handle_web_search(params: dict) -> dict[str, Any]:
    return _get_web_search().search(
        query=params["query"],
        max_results=params.get("max_results", 10),
    )


def _handle_web_fetch(params: dict) -> dict[str, Any]:
    return _get_web_search().fetch(
        url=params["url"],
        timeout=params.get("timeout", 15),
    )


TOOL_MAP: dict[str, Callable[[dict], dict[str, Any]]] = {
    "uia_get_interactive": _handle_uia_get_interactive,
    "uia_click":           _handle_uia_click,
    "uia_type":            _handle_uia_type,
    "uia_find":            _handle_uia_find,
    "uia_get_property":    _handle_uia_get_property,
    "uia_fingerprint":     _handle_uia_fingerprint,
    "uia_find_at_point":   _handle_uia_find_at_point,
    # Browser
    "web_launch":          _handle_web_launch,
    "web_connect_cdp":     _handle_web_connect_cdp,
    "web_launch_browser":  _handle_web_launch_browser,
    "web_navigate":        _handle_web_navigate,
    "web_get_interactive": _handle_web_get_interactive,
    "ext_get_recorded_events": _handle_ext_get_recorded_events,
    "ext_set_capture":       _handle_ext_set_capture,
    "web_click_selector":  _handle_web_click_selector,
    "web_click_role":      _handle_web_click_role,
    "web_fill":            _handle_web_fill,
    "web_type":            _handle_web_type,
    "web_scroll":          _handle_web_scroll,
    "web_close":           _handle_web_close,
    "web_start_recording":        _handle_web_start_recording,
    "web_stop_recording":         _handle_web_stop_recording,
    "web_get_recorded_events":    _handle_web_get_recorded_events,
    # Screenshot (mss)
    "screenshot_full":     _handle_screenshot_full,
    "screenshot_region":   _handle_screenshot_region,
    "screenshot_monitors": _handle_screenshot_monitors,
    # Global input listener (pynput)
    "global_listener_start": _handle_global_listener_start,
    "global_listener_stop":  _handle_global_listener_stop,
    "global_listener_poll":  _handle_global_listener_poll,
    # Event collector (unified events from global listener + extension)
    "event_collector_poll":  _handle_event_collector_poll,
    "event_collector_start": _handle_event_collector_start,
    "event_collector_stop":  _handle_event_collector_stop,
    # OCR (PaddleOCR)
    "ocr_recognize":       _handle_ocr_recognize,
    # Office document generators
    "word_generate":       _handle_word_generate,
    "excel_generate":      _handle_excel_generate,
    "ppt_generate":        _handle_ppt_generate,
    # Office COM automation (live editing)
    "office_detect":       _handle_office_detect,
    "word_com_read":       _handle_word_com_read,
    "word_com_edit":       _handle_word_com_edit,
    "excel_com_read":      _handle_excel_com_read,
    "excel_com_edit":      _handle_excel_com_edit,
    "ppt_com_read":        _handle_ppt_com_read,
    "ppt_com_edit":        _handle_ppt_com_edit,
    # Code sandbox (exec_python)
    "exec_python":         _handle_exec_python,
    "doc_code_exec":       _handle_doc_code_exec,
    "web_code_exec":       _handle_web_code_exec,
    # Web search (DuckDuckGo + page fetch)
    "web_search":          _handle_web_search,
    "web_fetch":           _handle_web_fetch,
    # Pre-warm (no-op, just confirms engine is alive)
    "prewarm":             lambda _params: {"status": "ready"},
    # Browser connection status (extension + browser engine)
    "browser_status":      lambda _params: _get_browser_status(),
}


def main() -> None:
    """Run the stdin/stdout event loop."""
    # Start extension WebSocket server on port 19840
    try:
        from engine.extension_ws import start_extension_ws_server
        start_extension_ws_server()
    except Exception as e:
        sys.stderr.write(f"[python-engine] Extension WS server failed to start: {e}\n")
        sys.stderr.flush()

    # Warmup: pre-initialize EasyOCR so the first OCR request doesn't
    # block the bridge Mutex for minutes (downloading ~200MB models +
    # loading PyTorch). A blocked Mutex piles up Tauri IPC calls and
    # triggers a WebView2 resource-request handler panic (0xc0000409).
    try:
        import sys
        sys.stderr.write("[python-engine] Warming up EasyOCR (first run downloads ~200MB models)...\n")
        sys.stderr.flush()
        _get_ocr().warmup()
        sys.stderr.write("[python-engine] EasyOCR warmup complete.\n")
        sys.stderr.flush()
    except Exception as e:
        sys.stderr.write(f"[python-engine] EasyOCR warmup failed (OCR unavailable): {e}\n")
        sys.stderr.flush()

    # Signal ready (Rust side reads this line to know the engine is alive)
    write_response({"id": "__ready__", "ok": True, "data": {"version": "0.1.0"}})

    while True:
        try:
            req = read_request()
        except Exception as e:
            write_response(fail("__parse__", f"JSON parse error: {e}"))
            continue

        if req is None:
            break  # stdin closed, shut down

        req_id = req.get("id", "")
        tool = req.get("tool", "")
        params = req.get("params", {})

        handler = TOOL_MAP.get(tool)
        if handler is None:
            write_response(fail(req_id, f"Unknown tool: {tool}"))
            continue

        try:
            data = handler(params)
            write_response(ok(req_id, data))
        except Exception as e:
            write_response(fail(req_id, traceback.format_exc()))


if __name__ == "__main__":
    main()
